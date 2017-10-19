import openpyxl

plik = openpyxl.load_workbook("example.xlsx")

arkusze = plik.get_sheet_names()

print(arkusze)

arkusz = plik.get_sheet_by_name("Owoce")

print(f"Aktywny arkusz: {arkusz.title}")

print(plik.active)

# komórki
komorka = arkusz['A1']
print(komorka)
print(komorka.value)

# koordynaty
print(f"Adres komórki: {komorka.coordinate}")
print(f"Kolumna komórki: {komorka.column}")
print(f"Wiersz komórki: {komorka.row}")

# komórka
print(arkusz.cell(row=3, column=2))

# rozmiar arkusza
print(arkusz.max_column)
print(arkusz.max_row)

# zamiana liter na cyfry i vice-versa
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(96))
print(get_column_letter(905))

print(column_index_from_string("AAG"))
print(column_index_from_string("AA"))

print(arkusz["A1":"C2"])


for wiersz in arkusz["A1":"C7"]:
    for kom in wiersz:
        print(f"Komorka {kom.coordinate} ma wartość {kom.value}")

arkusz["D3"] = "Moja wartość"

print(arkusz["D3"].value)

plik.save("example2.xlsx")
plik.close()








